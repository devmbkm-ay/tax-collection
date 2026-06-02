"""
PDF report generation, JSON backup and CSV export for WorldRemit transactions.
"""

import csv
import io
import json
from dataclasses import asdict
from datetime import datetime
from typing import Dict, List

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from .models import LOGO_PATH, TRANSLATIONS, WorldRemitTransaction


def _localise_date(date_str: str, months: Dict[str, str]) -> str:
    for en, local in months.items():
        date_str = date_str.replace(en, local)
    return date_str


def generate_pdf_report(
    transactions: List[WorldRemitTransaction],
    eur_rates: Dict[str, float],
    output_file: str = "worldremit_transactions_report.pdf",
    recipient_name: str = "",
    start_year: int = 2023,
    end_year: int = 2023,
    lang: str = 'fr',
    declarant_name: str = "",
    declarant_address: str = "",
) -> str:
    """
    Build and save a PDF tax report.
    Returns the output file path on success, raises on failure.
    """
    tr = TRANSLATIONS.get(lang, TRANSLATIONS['fr'])
    localise = lambda d: _localise_date(d, tr['months'])

    doc = SimpleDocTemplate(output_file, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()

    # ── Logo ────────────────────────────────────────────────────────────
    if LOGO_PATH.exists():
        logo = Image(str(LOGO_PATH), width=2.2 * inch, height=0.6 * inch)
        logo.hAlign = 'CENTER'
        story.append(logo)
        story.append(Spacer(1, 10))

    # ── Title ────────────────────────────────────────────────────────────
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        fontSize=18, spaceAfter=30,
        alignment=TA_CENTER, textColor=colors.darkblue,
    )
    story.append(Paragraph(tr['title'], title_style))

    # ── Declarant block (optional) ────────────────────────────────────────
    if declarant_name or declarant_address:
        decl_style = ParagraphStyle(
            'Declarant', parent=styles['Normal'],
            fontSize=10, spaceAfter=6, leading=15,
            leftIndent=20, borderPadding=(8, 12, 8, 12),
            backColor=colors.Color(0.93, 0.95, 1.0),
            borderColor=colors.darkblue, borderWidth=0.5, borderRadius=4,
        )
        decl_label = tr.get('declarant', 'Déclarant') if lang == 'fr' else 'Declarant'
        lines = [f"<b>{decl_label}</b><br/>"]
        if declarant_name:
            lines.append(f"{declarant_name}<br/>")
        if declarant_address:
            for addr_line in declarant_address.splitlines():
                if addr_line.strip():
                    lines.append(f"{addr_line.strip()}<br/>")
        story.append(Paragraph("".join(lines), decl_style))
        story.append(Spacer(1, 14))

    # ── Summary header ───────────────────────────────────────────────────
    currency_totals: Dict[str, float] = {}
    eur_total = 0.0
    for t in transactions:
        if t.amount != 'N/A':
            try:
                currency_totals[t.currency] = currency_totals.get(t.currency, 0) + float(t.amount)
            except ValueError:
                pass
        if t.amount_eur != 'N/A':
            try:
                eur_total += float(t.amount_eur)
            except ValueError:
                pass

    rate_parts = [
        f"{cur} : 1 EUR ≈ {eur_rates[cur]:.2f} {cur}"
        for cur in sorted(currency_totals)
        if cur in eur_rates
    ]
    rate_note = "  |  ".join(rate_parts) or "N/A"
    period_str = str(start_year) if start_year == end_year else f"{start_year} – {end_year}"
    report_date = localise(datetime.now().strftime("%d %B %Y"))

    info_style = ParagraphStyle(
        'Info', parent=styles['Normal'],
        fontSize=10, spaceAfter=20, alignment=TA_LEFT,
    )
    info_text = (
        f"<b>{tr['report_date']} :</b> {report_date}<br/>"
        f"<b>{tr['total_txn']} :</b> {len(transactions)}<br/>"
        f"<b>{tr['recipient']} :</b> {recipient_name}<br/>"
        f"<b>{tr['period']} :</b> {period_str}<br/>"
        f"<b>{tr['total_eur']} :</b> {eur_total:,.2f} EUR<br/>"
        f"<i>{tr['rates_note']} : {rate_note}</i>"
    )
    story.append(Paragraph(info_text, info_style))
    story.append(Spacer(1, 20))

    if not transactions:
        story.append(Paragraph(tr['no_txn'], styles['Normal']))
        doc.build(story)
        return output_file

    # ── Table ─────────────────────────────────────────────────────────────
    rows = [[
        tr['col_date'], tr['col_amount'], tr['col_currency'],
        tr['col_eur'],  tr['col_txn'],   tr['col_country'],
    ]]
    for t in sorted(transactions, key=lambda x: x.sort_key):
        rows.append([
            localise(t.date),
            t.amount,
            t.currency,
            t.amount_eur,
            t.transaction_number,
            t.country or '—',
        ])

    col_widths = [1.6*inch, 1.05*inch, 0.55*inch, 0.8*inch, 1.5*inch, 1.0*inch]
    table = Table(rows, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.darkblue),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.whitesmoke),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  9),
        ('BOTTOMPADDING', (0, 0), (-1, 0),  10),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 7.5),
        ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN',         (1, 1), (3, -1),  'RIGHT'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.Color(0.93, 0.95, 1.0)]),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(table)

    # ── Currency summary ──────────────────────────────────────────────────
    story.append(Spacer(1, 30))
    summary_style = ParagraphStyle(
        'Summary', parent=styles['Normal'], fontSize=10, leftIndent=20,
    )
    lines = [f"<b>{tr['summary_title']} :</b><br/>"]
    for cur, total in sorted(currency_totals.items()):
        rate = eur_rates.get(cur, 0)
        eur_eq = total / rate if rate > 0 else 0
        lines.append(f"{cur} : {total:,.2f}  (≈ {eur_eq:,.2f} EUR)<br/>")
    lines.append(f"<br/><b>{tr['total_eur_lbl']} : {eur_total:,.2f} EUR</b><br/>")
    lines.append(f"<i>{tr['disclaimer']}</i>")
    story.append(Paragraph("".join(lines), summary_style))

    doc.build(story)
    print(f"PDF report generated: {output_file}")
    return output_file


def save_json_backup(
    transactions: List[WorldRemitTransaction],
    output_file: str = "worldremit_transactions.json",
) -> str:
    """Serialise transactions to JSON. Returns the output file path."""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump([asdict(t) for t in transactions], f, indent=2, ensure_ascii=False)
    print(f"JSON backup saved: {output_file}")
    return output_file


def export_csv(
    transactions: List[WorldRemitTransaction],
    lang: str = 'fr',
) -> str:
    """
    Serialise transactions to a CSV string (UTF-8 with BOM for Excel compatibility).
    Returns the CSV content as a string.
    """
    from .models import TRANSLATIONS
    tr = TRANSLATIONS.get(lang, TRANSLATIONS['fr'])

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    writer.writerow([
        tr['col_date'], tr['col_amount'], tr['col_currency'],
        tr['col_eur'], tr['col_txn'], tr['col_country'],
    ])
    for t in sorted(transactions, key=lambda x: x.sort_key):
        writer.writerow([
            t.date, t.amount, t.currency,
            t.amount_eur, t.transaction_number, t.country or '—',
        ])

    # UTF-8 BOM so Excel opens it correctly without encoding issues
    return '﻿' + output.getvalue()


def export_xlsx(
    transactions: List[WorldRemitTransaction],
    recipient_name: str = "",
    period: str = "",
    lang: str = "fr",
) -> bytes:
    """Generate a styled .xlsx workbook and return it as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from .models import TRANSLATIONS

    tr = TRANSLATIONS.get(lang, TRANSLATIONS["fr"])
    wb = Workbook()
    ws = wb.active
    ws.title = period or "Transactions"

    # ── Palette ──────────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="2D2D2D")
    ALT_FILL    = PatternFill("solid", fgColor="F4F1EA")
    TOTAL_FILL  = PatternFill("solid", fgColor="E8E3D8")
    thin = Side(style="thin", color="DDDDDD")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title block ───────────────────────────────────────────────────────
    if recipient_name:
        ws.append([tr.get("report_title", "WorldRemit — Rapport fiscal"), "", "", "", "", ""])
        ws.append([recipient_name, "", "", "", "", ""])
        ws.append([period, "", "", "", "", ""])
        ws.append([])
        title_rows = 4
    else:
        title_rows = 0

    # ── Header row ────────────────────────────────────────────────────────
    headers = [tr["col_date"], tr["col_recipient"], tr["col_amount"],
               tr["col_currency"], tr["col_eur"], tr["col_txn"], tr["col_country"]]
    ws.append(headers)
    header_row = title_rows + 1

    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    # ── Data rows ─────────────────────────────────────────────────────────
    sorted_txns = sorted(transactions, key=lambda x: x.sort_key)
    total_eur = 0.0

    for i, t in enumerate(sorted_txns):
        try:
            eur_val = float(t.amount_eur.replace(",", ".")) if t.amount_eur and t.amount_eur != "N/A" else None
        except ValueError:
            eur_val = None
        if eur_val:
            total_eur += eur_val

        row = [t.date, t.recipient_name or "—", t.amount, t.currency,
               eur_val if eur_val is not None else t.amount_eur,
               t.transaction_number, t.country or "—"]
        ws.append(row)
        data_row = header_row + 1 + i
        fill = ALT_FILL if i % 2 == 1 else None
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col)
            if fill:
                cell.fill = fill
            cell.border = cell_border
            cell.font = Font(name="Calibri", size=10)
            if col in (3, 5):  # amount columns
                cell.alignment = Alignment(horizontal="right")

    # ── Total row ─────────────────────────────────────────────────────────
    total_row_idx = header_row + 1 + len(sorted_txns)
    total_label = "TOTAL" if lang == "en" else "TOTAL"
    ws.append(["", "", "", total_label, round(total_eur, 2), "", ""])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_idx, column=col)
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.border = cell_border
        if col == 5:
            cell.number_format = "#,##0.00"

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = [16, 24, 10, 10, 10, 22, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[header_row].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
